"""
routes/custom_vocab_routes.py
自訂單字本功能 — 支援上傳 Excel/CSV、線上編輯、進度書籤、以及匯出
"""
import json
import os
import io
from datetime import datetime, date

from flask import Blueprint, render_template, request, jsonify, current_app, send_file
from flask_login import login_required, current_user

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

custom_vocab_bp = Blueprint('custom_vocab', __name__, template_folder='../templates')

# ─── 常數 ─────────────────────────────────────────────────────────
CUSTOM_VOCAB_DIR = None   # 延遲初始化，在 request context 內使用 current_app.root_path

EXPECTED_COLUMNS = {
    'word':        ['word', 'english', 'english_word', '單字', '英文'],
    'definition':  ['definition', 'chinese', 'chinese_definition', 'meaning', '中文', '意思', '定義', '中文意思', '中文翻譯'],
    'pos':         ['pos', 'parts_of_speech', 'part_of_speech', '詞性'],
    'example_en':  ['example_en', 'example_english', 'english_example', '英文例句'],
    'example_zh':  ['example_zh', 'example_chinese', 'chinese_example', '中文例句'],
    'preposition': ['介係詞', 'preposition', 'prep'],
    'exam_tip':    ['出題重點', 'exam_tip', 'tip', 'notes', '備註'],
}


# ─── 輔助函數 ──────────────────────────────────────────────────────
def _user_vocab_path():
    """回傳當前使用者的自訂單字 JSON 路徑"""
    data_dir = os.path.join(current_app.root_path, 'static', 'data', 'custom_vocab')
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, f'user_{current_user.id}.json')


def _load_custom_vocab():
    """載入當前使用者的自訂單字本"""
    path = _user_vocab_path()
    if not os.path.exists(path):
        return {'list_name': '我的單字本', 'bookmark': None, 'updated_at': None, 'words': []}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _save_custom_vocab(data):
    """儲存當前使用者的自訂單字本"""
    data['updated_at'] = datetime.utcnow().isoformat()
    path = _user_vocab_path()
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _normalize_col(col_name: str, mapping: dict) -> str | None:
    """根據 EXPECTED_COLUMNS 對應表，把 Excel 欄位名稱對應到系統標準欄位名稱"""
    col_lower = col_name.strip().lower()
    for key, aliases in mapping.items():
        if col_lower in [a.lower() for a in aliases]:
            return key
    return None


def _parse_excel_bytes(file_bytes: bytes) -> list[dict]:
    """解析 Excel bytes，回傳 word 物件列表
    
    支援兩種格式：
    1. 標準格式：標題列含 Word, Definition, POS, Example_EN, Example_ZH
    2. 使用者自訂格式（vocabulary.xlsx 格式）：
       A=單字, B=詞性, C=中文意思（無標題）, D=介係詞, E=出題重點
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 第一列為標題
    raw_headers = [str(h).strip() if h else '' for h in rows[0]]
    col_map = {}   # {system_key: col_index}
    for i, h in enumerate(raw_headers):
        if not h:
            continue
        key = _normalize_col(h, EXPECTED_COLUMNS)
        if key:
            col_map[key] = i

    # ─── 位置式回退（適用於使用者的 vocabulary.xlsx 格式）─────────────
    # 判斷依據：標題列第 0 欄是「單字」（已對應到 word），
    # 但第 2 欄標題是空的（None），需依位置推斷它是「中文意思」
    if 'word' in col_map and 'definition' not in col_map:
        # 嘗試找第一個非空、非 word、非 pos 的欄位 → 視為 definition
        used_cols = set(col_map.values())
        for i, h in enumerate(raw_headers):
            if i not in used_cols and i != col_map.get('word') and i != col_map.get('pos'):
                # 檢查第一筆資料該欄是否有值且看起來像中文
                sample = rows[1][i] if len(rows) > 1 else None
                if sample and any('\u4e00' <= c <= '\u9fff' for c in str(sample)):
                    col_map['definition'] = i
                    break

    # 抓「出題重點」欄（用來當作 exam_tip，合併進 examples）
    for i, h in enumerate(raw_headers):
        if h and _normalize_col(h, {'exam_tip': ['出題重點', 'exam_tip', 'tip', 'notes', '備註']}):
            col_map['exam_tip'] = i
        if h and _normalize_col(h, {'preposition': ['介係詞', 'preposition', 'prep']}):
            col_map['preposition'] = i

    words = []
    for row in rows[1:]:
        word_val = str(row[col_map['word']]).strip() if 'word' in col_map else ''
        if not word_val or word_val.lower() == 'none':
            continue

        def _get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ''
            val = row[idx]
            return str(val).strip() if val else ''

        pos_raw = _get('pos')
        pos_list = [p.strip() for p in pos_raw.replace('/', ',').split(',') if p.strip()] if pos_raw else []

        # 組合 definition：中文意思 + 介係詞補充
        definition = _get('definition')
        prep = _get('preposition')
        if prep and prep.lower() != 'none':
            definition = f"{definition}（{prep}）".strip('（）').replace('（）', '') if not definition else f"{definition}（{prep}）"

        # 把「出題重點」存進 examples（當作英文例句提示）
        examples = []
        en = _get('example_en')
        zh = _get('example_zh')
        exam_tip = _get('exam_tip')
        if en or zh:
            examples.append({'english': en, 'chinese': zh})
        elif exam_tip and exam_tip.lower() != 'none':
            examples.append({'english': exam_tip, 'chinese': ''})

        words.append({
            'word':             word_val,
            'definition':       definition,
            'parts_of_speech':  pos_list,
            'examples':         examples,
        })

    wb.close()
    return words


# ─── 頁面路由 ──────────────────────────────────────────────────────
@custom_vocab_bp.route('/custom')
@login_required
def custom_editor():
    """自訂單字本編輯器頁面"""
    return render_template('vocab/custom.html')


# ─── API ──────────────────────────────────────────────────────────
@custom_vocab_bp.route('/api/custom/words', methods=['GET'])
@login_required
def api_custom_get():
    """GET 取得自訂單字本"""
    data = _load_custom_vocab()
    return jsonify(data)


@custom_vocab_bp.route('/api/custom/words', methods=['POST'])
@login_required
def api_custom_save():
    """POST 整批儲存單字清單（來自前端編輯器）"""
    body = request.get_json()
    if body is None:
        return jsonify({'ok': False, 'msg': '無效的 JSON'}), 400

    data = _load_custom_vocab()
    data['list_name'] = body.get('list_name', data.get('list_name', '我的單字本'))
    data['words'] = body.get('words', [])
    _save_custom_vocab(data)
    return jsonify({'ok': True, 'count': len(data['words'])})


@custom_vocab_bp.route('/api/custom/bookmark', methods=['POST'])
@login_required
def api_custom_bookmark():
    """POST 設定/清除書籤（記錄今天編輯到哪個 index）"""
    body = request.get_json()
    data = _load_custom_vocab()
    idx = body.get('index')   # None = 清除書籤
    if idx is not None:
        data['bookmark'] = {
            'index': int(idx),
            'date': date.today().isoformat(),
            'word': (data['words'][int(idx)]['word'] if int(idx) < len(data['words']) else '')
        }
    else:
        data['bookmark'] = None
    _save_custom_vocab(data)
    return jsonify({'ok': True, 'bookmark': data.get('bookmark')})


@custom_vocab_bp.route('/api/custom/upload', methods=['POST'])
@login_required
def api_custom_upload():
    """POST 上傳 Excel 並解析成單字清單"""
    if not HAS_OPENPYXL:
        return jsonify({'ok': False, 'msg': 'openpyxl 未安裝'}), 500

    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'msg': '未選擇檔案'}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'ok': False, 'msg': '僅支援 .xlsx 格式'}), 400

    try:
        words = _parse_excel_bytes(f.read())
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'解析失敗：{e}'}), 400

    return jsonify({'ok': True, 'words': words, 'count': len(words)})


@custom_vocab_bp.route('/api/custom/export', methods=['GET'])
@login_required
def api_custom_export():
    """GET 將自訂單字本匯出為 Excel"""
    if not HAS_OPENPYXL:
        return jsonify({'ok': False, 'msg': 'openpyxl 未安裝'}), 500

    data = _load_custom_vocab()
    words = data.get('words', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vocabulary'

    # Header
    ws.append(['Word', 'Definition', 'POS', 'Example_EN', 'Example_ZH'])

    for w in words:
        pos = ', '.join(w.get('parts_of_speech') or [])
        examples = w.get('examples') or []
        ex_en = examples[0].get('english', '') if examples else ''
        ex_zh = examples[0].get('chinese', '') if examples else ''
        ws.append([w.get('word', ''), w.get('definition', ''), pos, ex_en, ex_zh])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    list_name = data.get('list_name', 'vocabulary')
    safe_name = ''.join(c for c in list_name if c.isalnum() or c in ' _-').strip() or 'vocabulary'
    filename = f'{safe_name}.xlsx'

    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@custom_vocab_bp.route('/api/custom/template', methods=['GET'])
@login_required
def api_custom_template():
    """GET 下載空白範本 Excel"""
    if not HAS_OPENPYXL:
        return jsonify({'ok': False, 'msg': 'openpyxl 未安裝'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vocabulary'

    headers = ['Word', 'Definition', 'POS', 'Example_EN', 'Example_ZH']
    ws.append(headers)

    # 範例一列
    ws.append(['implement', '實作、執行', 'v.', 'We will implement the new feature.', '我們將實作新功能。'])

    # 設定欄寬
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 30, 10, 45, 45]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, as_attachment=True, download_name='vocab_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
